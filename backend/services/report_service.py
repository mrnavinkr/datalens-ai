"""
report_service.py — Section 21: Report Export (PDF, Excel, CSV).

Builds the "Final Data Intelligence Report" (Section 20) content and
renders it into a downloadable file. All figures come from the already-
persisted DatasetAnalysis / DatasetColumn rows — nothing is recomputed
or invented here.
"""
from __future__ import annotations

import csv
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from models import Dataset, DatasetAnalysis, DatasetColumn


def _report_dir(dataset_id: str, upload_dir: str) -> str:
    path = os.path.join(upload_dir, "..", "reports", dataset_id)
    path = os.path.normpath(path)
    os.makedirs(path, exist_ok=True)
    return path


def generate_pdf_report(
    dataset: Dataset,
    analysis: DatasetAnalysis,
    columns: list[DatasetColumn],
    upload_dir: str,
) -> str:
    out_dir = _report_dir(dataset.id, upload_dir)
    out_path = os.path.join(out_dir, "report.pdf")

    doc = SimpleDocTemplate(out_path, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Title"], textColor=colors.HexColor("#0B1220"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#14B8A6"), spaceBefore=14)
    body = styles["BodyText"]

    story = [
        Paragraph("DataLens AI — Data Intelligence Report", title_style),
        Paragraph(f"Dataset: {dataset.display_name}", body),
        Spacer(1, 12),

        Paragraph("Dataset Summary", h2),
        Table(
            [
                ["Rows", f"{dataset.total_rows:,}" if dataset.total_rows else "—"],
                ["Columns", str(dataset.total_columns or "—")],
                ["Data Quality", f"{round(analysis.quality_score)} / 100"],
                ["Data Usability", f"{round(analysis.usability_score)} / 100"],
                ["Status", analysis.usability_status],
            ],
            colWidths=[6 * cm, 8 * cm],
            style=TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]),
        ),

        Paragraph("Key Findings", h2),
        *[Paragraph(f"{i+1}. {f}", body) for i, f in enumerate(analysis.key_findings or [])],

        Paragraph("Recommended Actions", h2),
        *[Paragraph(a, body) for a in (analysis.recommended_actions or [])],

        Paragraph("Column Statistics", h2),
        Table(
            [["Column", "Type", "Null %", "Unique %"]] + [
                [c.name, c.data_type, f"{c.null_percentage}%", f"{c.unique_percentage}%"]
                for c in columns
            ],
            colWidths=[5 * cm, 3 * cm, 3 * cm, 3 * cm],
            style=TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B1220")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]),
        ),

        Spacer(1, 20),
        Paragraph("Overall Recommendation", h2),
        Paragraph(
            "Dataset is useful for further analysis. Perform the recommended data-quality "
            "checks above before proceeding.", body
        ),
        Spacer(1, 30),
        Paragraph("Developed by Er. Navin Kumar", styles["Normal"]),
    ]

    doc.build(story)
    return out_path


def generate_excel_report(
    dataset: Dataset,
    analysis: DatasetAnalysis,
    columns: list[DatasetColumn],
    upload_dir: str,
) -> str:
    out_dir = _report_dir(dataset.id, upload_dir)
    out_path = os.path.join(out_dir, "report.xlsx")

    wb = Workbook()
    header_fill = PatternFill(start_color="0B1220", end_color="0B1220", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in [
        ["Dataset", dataset.display_name],
        ["Rows", dataset.total_rows],
        ["Columns", dataset.total_columns],
        ["Data Quality Score", analysis.quality_score],
        ["Data Usability Score", analysis.usability_score],
        ["Usability Status", analysis.usability_status],
        ["Missing %", analysis.missing_percentage],
        ["Duplicate %", analysis.duplicate_percentage],
    ]:
        ws_summary.append(row)
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 30

    ws_cols = wb.create_sheet("Columns")
    headers = ["Column", "Type", "Null Count", "Null %", "Unique Count", "Unique %", "Mean", "Std Dev", "Outliers"]
    ws_cols.append(headers)
    for cell in ws_cols[1]:
        cell.fill = header_fill
        cell.font = header_font
    for c in columns:
        ws_cols.append([
            c.name, c.data_type, c.null_count, c.null_percentage,
            c.unique_count, c.unique_percentage, c.mean_value, c.std_value, c.outlier_count,
        ])
    for col_letter in "ABCDEFGHI":
        ws_cols.column_dimensions[col_letter].width = 16

    ws_findings = wb.create_sheet("Findings & Actions")
    ws_findings.append(["Key Findings"])
    ws_findings["A1"].font = header_font
    ws_findings["A1"].fill = header_fill
    for f in (analysis.key_findings or []):
        ws_findings.append([f])
    ws_findings.append([])
    ws_findings.append(["Recommended Actions"])
    for a in (analysis.recommended_actions or []):
        ws_findings.append([a])
    ws_findings.column_dimensions["A"].width = 70

    wb.save(out_path)
    return out_path


def generate_csv_report(
    dataset: Dataset,
    analysis: DatasetAnalysis,
    columns: list[DatasetColumn],
    upload_dir: str,
) -> str:
    out_dir = _report_dir(dataset.id, upload_dir)
    out_path = os.path.join(out_dir, "analysis_summary.csv")

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Dataset", dataset.display_name])
        writer.writerow(["Rows", dataset.total_rows])
        writer.writerow(["Columns", dataset.total_columns])
        writer.writerow(["Data Quality Score", analysis.quality_score])
        writer.writerow(["Data Usability Score", analysis.usability_score])
        writer.writerow(["Usability Status", analysis.usability_status])
        writer.writerow([])
        writer.writerow(["Column", "Type", "Null %", "Unique %", "Mean", "Std Dev", "Outliers"])
        for c in columns:
            writer.writerow([c.name, c.data_type, c.null_percentage, c.unique_percentage, c.mean_value, c.std_value, c.outlier_count])

    return out_path
